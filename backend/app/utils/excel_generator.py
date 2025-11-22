from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel files from extracted invoice data."""

    def __init__(self):
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_invoice_report(self, invoices: List[Dict], output_path: str) -> str:
        """
        Generate an Excel file with invoice data and summary.

        Args:
            invoices: List of extracted invoice dictionaries
            output_path: Path where to save the Excel file

        Returns:
            Path to the generated Excel file
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoices"

            # Create header row
            headers = [
                "Vendor Name",
                "Invoice Number",
                "Invoice Date",
                "Total Amount",
                "Line Items Count",
                "Flagged"
            ]

            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = self.thin_border

            # Add invoice data
            current_row = 2
            for invoice in invoices:
                ws.cell(row=current_row, column=1).value = invoice.get('vendor_name', '')
                ws.cell(row=current_row, column=2).value = invoice.get('invoice_number', '')
                ws.cell(row=current_row, column=3).value = invoice.get('invoice_date', '')

                total_amount = float(invoice.get('total_amount', 0))
                ws.cell(row=current_row, column=4).value = total_amount
                ws.cell(row=current_row, column=4).number_format = '$#,##0.00'

                line_items = invoice.get('line_items', [])
                ws.cell(row=current_row, column=5).value = len(line_items)

                flagged = invoice.get('flagged', False)
                ws.cell(row=current_row, column=6).value = "Yes" if flagged else "No"

                # Apply formatting to data rows
                for col_num in range(1, 7):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                    # Highlight flagged invoices
                    if flagged and col_num == 6:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

                current_row += 1

            # Add summary section
            summary_row = current_row + 2
            ws.cell(row=summary_row, column=1).value = "SUMMARY"
            ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)

            # Summary data
            total_invoices = len(invoices)
            total_amount = sum(float(inv.get('total_amount', 0)) for inv in invoices)
            flagged_count = sum(1 for inv in invoices if inv.get('flagged', False))

            summary_row += 1
            ws.cell(row=summary_row, column=1).value = "Total Invoices Processed:"
            ws.cell(row=summary_row, column=2).value = total_invoices
            ws.cell(row=summary_row, column=2).font = Font(bold=True)

            summary_row += 1
            ws.cell(row=summary_row, column=1).value = "Total Amount:"
            ws.cell(row=summary_row, column=2).value = total_amount
            ws.cell(row=summary_row, column=2).number_format = '$#,##0.00'
            ws.cell(row=summary_row, column=2).font = Font(bold=True)

            summary_row += 1
            ws.cell(row=summary_row, column=1).value = "Flagged Invoices (>$5,000):"
            ws.cell(row=summary_row, column=2).value = flagged_count
            ws.cell(row=summary_row, column=2).font = Font(bold=True)

            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 12

            wb.save(output_path)
            logger.info(f"Excel report generated at {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error generating Excel file: {str(e)}")
            raise

    def get_summary_data(self, invoices: List[Dict]) -> Dict:
        """Calculate summary statistics for invoices."""
        total_invoices = len(invoices)
        total_amount = sum(float(inv.get('total_amount', 0)) for inv in invoices)
        flagged_count = sum(1 for inv in invoices if inv.get('flagged', False))

        return {
            "total_invoices_processed": total_invoices,
            "total_amount": round(total_amount, 2),
            "flagged_invoices_count": flagged_count
        }
